import openpyxl
import re
import tkinter as tk
from tkinter import filedialog
import os

def choose_files():
    # 创建一个隐藏的 Tkinter 根窗口
    root = tk.Tk()
    root.withdraw()  # 隐藏根窗口

    # 固定的 Markdown 文件路径
    markdown_file = r"D:\Workspace\Stable\Python\单词整理\词组temp.md"

    # 选择输出的 Excel 文件的默认路径
    default_path = r"D:\Workspace\WPS"

    # 选择输出的 Excel 文件
    excel_file = filedialog.askopenfilename(
        title="选择输出的 Excel 文件",
        filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*")),
        initialdir=default_path  # 设置默认路径
    )
    if not excel_file:  # 如果用户没有选择文件，返回
        print("没有选择 Excel 文件")
        return None, None

    # 返回绝对路径
    return os.path.abspath(markdown_file), os.path.abspath(excel_file)


def load_markdown(file_path):
    """读取Markdown文件并提取单词和其后无序列表的内容"""
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    # 正则表达式匹配单词及其后面紧接的无序列表部分
    # 匹配每个单词后跟随的无序列表（如：* item1, * item2 等）
    pattern = r'\d+\.\s*\*\*([\s\S]*?)\*\*([\s\S]*?)\n\n'


    # 查找所有匹配的内容
    matches = re.findall(pattern, content, re.DOTALL)
    
    word_data = {}

    for match in matches:
        word = match[0].lower()    # 第一个匹配项是单词
        list_items = match[1].strip().split('\n')  # 第二个匹配项是无序列表的内容
        # 处理每一行
        cleaned_items = []
        for item in list_items:
            # 清除行前后的空格
            item = item.strip()
            
            # 去掉行前的 - 和空格（即去掉 Markdown 项符号）
            item = re.sub(r'^\s*-+\s*', '', item)  # 去掉开头的横杠和空格
            
            # 去掉所有星号（用于去除 Markdown 加粗、斜体符号）
            item = re.sub(r'\*', '', item)
            
            # 将处理过的行添加到结果列表
            cleaned_items.append(item)
        word_data[word] = cleaned_items
        
    return word_data




def update_excel(excel_file, word_data):
    """根据词汇数据更新Excel文件"""
    # 打开Excel文件
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    # 固定行高
    fixed_row_height = 13.5  # 可以根据需要调整行高

    # 去除第一列的重复项
    unique_words = set()
    rows_to_delete = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1), start=2):
        word_cell = row[0]
        word = word_cell.value.strip().lower()
        if word in unique_words:
            rows_to_delete.append(row_num)
        else:
            unique_words.add(word)
    # 从下往上删除行，避免索引混乱
    for row_num in reversed(rows_to_delete):
        sheet.delete_rows(row_num)

    # 遍历Excel表格的每一行（假设单词在第一列）
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):  # 从第二行开始
        word_cell = row[0]
        word = word_cell.value.strip()

        if word.lower() in word_data:
            # 获取该单词的释义和无序列表
            definitions = word_data[word.lower()]
            combined_definition = "\n".join(definitions)

            # 将定义粘贴到第二列单元格
            sheet.cell(row=word_cell.row, column=2, value=combined_definition)

        # 设置行高
        sheet.row_dimensions[word_cell.row].height = fixed_row_height

    # 保存修改后的Excel文件
    wb.save(excel_file)

def main():
    markdown_file,excel_file = choose_files()

    if markdown_file and excel_file:
        print(f"选择的 MD 文件: {markdown_file}")
        print(f"选择的 Excel 文件: {excel_file}")
        
    else:
        print("文件选择失败")

    # 加载Markdown文件中的单词数据
    word_data = load_markdown(markdown_file)
    print(word_data)
    # 更新Excel文件
    update_excel(excel_file, word_data)

if __name__ == "__main__":
    main()

import re
import openpyxl
from openpyxl.styles import Font, Alignment
import os

# --- 配置区：请在此处修改您的文件路径 ---
MARKDOWN_FILE_PATH = "D:\Workspace\Stable\Python\ReWordOrganizer\词根temp.md"  # << 修改为您实际的 Markdown 文件名和路径
# 默认创建新 Excel 文件
NEW_EXCEL_FILE_PATH = "parsed_roots_output.xlsx" # << 修改为您希望生成的新 Excel 文件名和路径
# -----------------------------------------

def parse_markdown_to_roots_data(markdown_content):
    """
    使用正则表达式解析 Markdown 文本内容，提取词根数据。
    预期 Markdown 格式:
    序号.  **主词根**
        * 语言：语言信息
        * 释义：释义信息
        * 备注：备注信息 (可选)
        * 词根形式：词根形式信息 (可选)
    """
    extracted_data = []
    # 正则表达式针对新的 Markdown 格式 (包含可选的“备注”行和可选的“词根形式”行)
    pattern = re.compile(
        r"^\s*\d+\.\s+\*\*(?P<root>.+?)\*\*\s*?\n"          # 1. 主词根 (加粗)
        r"\s*\*\s*语言：\s*(?P<language>.+?)\s*?\n"        # 2. 语言
        r"\s*\*\s*释义：\s*(?P<meaning>.+?)\s*?\n"          # 3. 释义
        r"(?:\s*\*\s*备注：\s*(?P<remarks>.*?)\s*?\n)?"      # 4. 备注 (可选行，内容可为空)
        r"(?:\s*\*\s*词根形式：\s*(?P<forms>.*?)\s*?\n)?"  # 5. 词根形式 (可选行，内容可为空)
        , re.MULTILINE
    )

    for match in pattern.finditer(markdown_content):
        data = match.groupdict()
        root = data.get("root", "").strip()
        language = data.get("language", "").strip()
        meaning = data.get("meaning", "").strip()
        # 处理可选组可能为 None 或仅包含空白的情况
        remarks = data.get("remarks")
        remarks_cleaned = remarks.strip() if remarks else ""
        
        forms = data.get("forms")
        forms_cleaned = forms.strip() if forms else ""

        if root: # 确保至少捕获到了词根
            extracted_data.append({
                "root": root,
                "language": language,
                "meaning": meaning,
                "remarks": remarks_cleaned,
                "forms": forms_cleaned
            })
            
    return extracted_data

def create_new_excel_from_data(data_list, excel_filename):
    """
    使用 openpyxl 将数据列表创建为一个新的 Excel 文件。
    输出3列：“词根”，“综合释义”，“词根形式”。
    """
    if not data_list:
        print("没有数据可写入新的 Excel 文件。")
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "词根数据"

    # 写入表头并设置样式
    headers = ["词根", "综合释义", "词根形式"]
    sheet.append(headers)
    for cell in sheet[1]: 
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据行
    for item in data_list:
        # 构建“综合释义”列
        meaning_parts = []
        if item.get("meaning"):
            meaning_parts.append(item.get("meaning"))
        if item.get("language"):
            meaning_parts.append(f"来自{item.get('language')}")
        if item.get("remarks"): # 只有当备注不为空时才添加
            meaning_parts.append(item.get("remarks"))
        
        combined_meaning = ", ".join(filter(None, meaning_parts)) # filter(None,...) 移除空字符串

        row_data = [
            item.get("root", ""),
            combined_meaning,
            item.get("forms", "")
        ]
        sheet.append(row_data)

    # 调整列宽
    column_letters = ['A', 'B', 'C']
    for col_idx, column_letter in enumerate(column_letters):
        max_length = 0
        # 计算表头长度
        if len(headers[col_idx]) > max_length:
            max_length = len(headers[col_idx])
        # 计算数据列长度
        for i in range(2, sheet.max_row + 1): # 从第二行开始（数据行）
            cell_value = sheet.cell(row=i, column=col_idx + 1).value
            if cell_value:
                # 对“综合释义”列，字符可能较多，可以考虑设置一个最大宽度上限或根据实际情况调整
                current_len = len(str(cell_value))
                if column_letter == 'B': # "综合释义"列
                    current_len = min(current_len, 60) # 例如，综合释义列最大宽度限制在60个字符左右
                if current_len > max_length:
                    max_length = current_len
        
        adjusted_width = (max_length + 3) if max_length > 0 else 20 # 最小宽度20
        sheet.column_dimensions[column_letter].width = adjusted_width
        # 对于“综合释义”列，启用自动换行
        if column_letter == 'B':
             for i in range(1, sheet.max_row + 1):
                sheet.cell(row=i, column=col_idx + 1).alignment = Alignment(wrap_text=True, vertical="top")


    # 保存文件
    try:
        workbook.save(excel_filename)
        print(f"成功将 {len(data_list)} 条数据写入到新的 Excel 文件 '{excel_filename}'")
    except PermissionError:
        print(f"保存 Excel 文件 '{excel_filename}' 失败：权限不足。可能文件已被打开或您没有写入权限。")
    except Exception as e:
        print(f"保存 Excel 文件 '{excel_filename}' 时发生错误: {e}")

def main():
    # 检查 Markdown 文件是否存在
    if not os.path.exists(MARKDOWN_FILE_PATH):
        print(f"错误: Markdown 文件 '{MARKDOWN_FILE_PATH}' 未找到。请在脚本顶部的配置区设置正确的文件路径。")
        return

    try:
        with open(MARKDOWN_FILE_PATH, 'r', encoding='utf-8') as f:
            markdown_content = f.read()
    except Exception as e:
        print(f"读取 Markdown 文件 '{MARKDOWN_FILE_PATH}' 时发生错误: {e}")
        return

    roots_data = parse_markdown_to_roots_data(markdown_content)

    if roots_data:
        print(f"从 Markdown 文件中成功解析出 {len(roots_data)} 条词根数据。")
        create_new_excel_from_data(roots_data, NEW_EXCEL_FILE_PATH)
    else:
        print(f"未能从 '{MARKDOWN_FILE_PATH}' 中解析到任何词根数据。请检查：")
        print("1. 文件路径是否正确。")
        print("2. 文件内容是否符合预期的 Markdown 格式 (包含序号、**词根**、* 语言、* 释义等)。")
        print("3. 正则表达式是否能匹配您的 Markdown 格式。")

if __name__ == "__main__":
    # 关于追加到现有文件的说明：
    # openpyxl 可以通过 openpyxl.load_workbook("现有文件.xlsx") 加载现有文件，
    # 然后获取工作表，找到最后一行，再 sheet.append(新数据行)，最后 workbook.save("现有文件.xlsx")。
    # 但如您所担心的，这涉及到加载和重写整个文件，如果原始文件样式非常复杂或特殊，
    # 存在微小的样式变动风险。因此，本脚本默认创建新文件以确保安全。
    # 如果您确实需要追加功能，可以基于以上提示进行修改，但请务必先在副本上测试。
    main()
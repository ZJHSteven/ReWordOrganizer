import re
from openpyxl import Workbook, load_workbook

# 读取 Markdown 文件内容
with open('d:/Workspace/Stable/Python/单词整理/词组temp.md', 'r', encoding='utf-8') as file:
    content = file.read()



# 定义正则表达式模式来匹配词根信息，处理带加号的情况
pattern = r'词根："(.*?)"（(.*?)，来自(.*?)“(.*?)(?:\+(.*?))?”）'

# 查找所有匹配的内容
matches = re.findall(pattern, content)

# 存储词根信息的字典，用于去重
root_dict = {}

# 遍历匹配结果
for match in matches:
    root = match[0]
    meaning = match[1]
    language = match[2]
    original_word = match[3]
    if match[4]:
        original_word += f"+{match[4]}"

    # 检查词根是否已经存在于字典中
    if root not in root_dict:
        root_dict[root] = {
            'meaning': meaning,
            'language': language,
            'original_word': original_word
        }

# 打开词根文件
try:
    wb_root = load_workbook(r"D:\Workspace\WPS\词根.xlsx")
    ws_root = wb_root.active
    # 获取词根文件第一列的所有值
    root_column_values = [cell.value for cell in ws_root['A'] if cell.value]
except FileNotFoundError:
    print("未找到词根文件，跳过重复检查。")
    root_column_values = []

# 过滤掉重复的词根和原词
filtered_root_dict = {}
for root, info in root_dict.items():
    original_word = info['original_word']
    if root not in root_column_values and original_word not in root_column_values:
        filtered_root_dict[root] = info

root_dict = filtered_root_dict

# 创建一个新的 Excel 工作簿
wb = Workbook()
ws = wb.active

# 设置表头
ws.append(['原词', '含义', '', '词根'])

# 填充数据到 Excel 表格
for root, info in root_dict.items():
    # 在含义前面添加带引号的词根
    meaning_with_language = f'"{root}", {info["meaning"]}（{info["language"]}）'
    ws.append([info['original_word'], meaning_with_language, '', root])

# 保存 Excel 文件
wb.save('d:/Workspace/Stable/Python/单词整理/词根整理.xlsx')

# 输出词根信息
print("词根\t语言来源\t原意")
for root, info in root_dict.items():
    print(f"{root}\t{info['language']}\t{info['original_word']} ({info['meaning']})")

